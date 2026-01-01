import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from datetime import datetime
import os
import numpy as np
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.fonts import addMapping
import warnings
warnings.filterwarnings('ignore')

class SalesAutomationSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("ሰናይት Sales Automation System")
        self.root.geometry("1100x750")
        
        # Variables
        self.file_path = None
        self.df = None
        self.items_list = []
        self.current_item = None
        self.shop_owner = "ቀጠና 5 ሰናይት"  # Default shop owner
        self.daily_sales = []  # Track sales for PDF export
        
        # Colors
        self.bg_color = "#f0f0f0"
        self.button_color = "#4CAF50"
        self.entry_bg = "white"
        
        # Font paths
        self.amharic_font_path = None
        self.find_amharic_font()
        
        self.root.configure(bg=self.bg_color)
        
        # Create UI
        self.create_widgets()
    
    def find_amharic_font(self):
        """Find Amharic font on system"""
        possible_fonts = [
            "Yebs.ttf",
            "Yebs.ttf",
            "Yebs.ttf",
            "Yebs.ttf",
            "Yebs.ttf"
        ]
        
        for font_path in possible_fonts:
            if os.path.exists(font_path):
                self.amharic_font_path = font_path
                print(f"Found Amharic font at: {font_path}")
                return
        
        # If not found, try to download or use default
        self.amharic_font_path = None
        print("Warning: Amharic font not found. PDFs may not display Amharic correctly.")
    
    def register_amharic_font(self):
        """Register Amharic font for PDF generation"""
        if self.amharic_font_path and os.path.exists(self.amharic_font_path):
            try:
                # Register the font with a name
                pdfmetrics.registerFont(TTFont('Amharic', self.amharic_font_path))
                pdfmetrics.registerFont(TTFont('Amharic-Bold', self.amharic_font_path))
                
                # Map the font
                addMapping('Amharic', 0, 0, 'Amharic')
                addMapping('Amharic', 1, 0, 'Amharic-Bold')
                return True
            except Exception as e:
                print(f"Could not register Amharic font: {e}")
                return False
        return False
    
    def create_widgets(self):
        # Title
        title_frame = tk.Frame(self.root, bg="#2E8B57")
        title_frame.pack(fill=tk.X, pady=(0, 10))
        
        title_label = tk.Label(title_frame, text="ሰናይት Sales Automation System", 
                               font=("Arial", 16, "bold"), 
                               fg="white", bg="#2E8B57")
        title_label.pack(pady=10)
        
        # Main container
        main_container = tk.Frame(self.root, bg=self.bg_color)
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Left panel - File and Item selection
        left_panel = tk.Frame(main_container, bg=self.bg_color)
        left_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        # Middle panel - Sales entry
        middle_panel = tk.Frame(main_container, bg=self.bg_color)
        middle_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        
        # Right panel - PDF and Shop info
        right_panel = tk.Frame(main_container, bg=self.bg_color)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        # ========== LEFT PANEL ==========
        
        # Shop Owner section
        shop_section = tk.LabelFrame(left_panel, text="Shop Information", 
                                    font=("Arial", 10, "bold"),
                                    bg=self.bg_color, padx=10, pady=10)
        shop_section.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(shop_section, text="Shop Owner Name:", bg=self.bg_color).pack(anchor="w", pady=(0, 5))
        self.shop_owner_entry = tk.Entry(shop_section, width=30, bg=self.entry_bg)
        self.shop_owner_entry.insert(0, self.shop_owner)
        self.shop_owner_entry.pack(fill=tk.X, pady=(0, 10))
        
        # Font status
        font_status_frame = tk.Frame(shop_section, bg=self.bg_color)
        font_status_frame.pack(fill=tk.X, pady=(0, 5))
        
        font_status = "✅ Amharic font available" if self.amharic_font_path else "⚠️ Amharic font not found"
        font_color = "green" if self.amharic_font_path else "orange"
        tk.Label(font_status_frame, text=font_status, bg=self.bg_color, 
                fg=font_color, font=("Arial", 8)).pack(anchor="w")
        
        # File selection section
        file_section = tk.LabelFrame(left_panel, text="File Operations", 
                                    font=("Arial", 10, "bold"),
                                    bg=self.bg_color, padx=10, pady=10)
        file_section.pack(fill=tk.X, pady=(0, 10))
        
        # File path display
        file_path_frame = tk.Frame(file_section, bg=self.bg_color)
        file_path_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(file_path_frame, text="Excel File:", bg=self.bg_color).pack(side=tk.LEFT)
        self.file_label = tk.Label(file_path_frame, text="No file loaded", 
                                  fg="gray", bg=self.bg_color, anchor="w")
        self.file_label.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))
        
        # Buttons frame
        button_frame = tk.Frame(file_section, bg=self.bg_color)
        button_frame.pack(fill=tk.X, pady=5)
        
        tk.Button(button_frame, text="Load Excel", 
                 command=self.load_excel_file,
                 bg=self.button_color, fg="white",
                 width=15).pack(side=tk.LEFT, padx=(0, 5))
        
        tk.Button(button_frame, text="Create New Excel", 
                 command=self.create_new_excel,
                 bg="#2196F3", fg="white",
                 width=15).pack(side=tk.LEFT)
        
        # Item selection section
        item_section = tk.LabelFrame(left_panel, text="Item Selection", 
                                    font=("Arial", 10, "bold"),
                                    bg=self.bg_color, padx=10, pady=10)
        item_section.pack(fill=tk.BOTH, expand=True)
        
        # Search box
        search_frame = tk.Frame(item_section, bg=self.bg_color)
        search_frame.pack(fill=tk.X, pady=(0, 5))
        
        tk.Label(search_frame, text="Search:", bg=self.bg_color).pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(search_frame, textvariable=self.search_var, 
                                    width=30, bg=self.entry_bg)
        self.search_entry.pack(side=tk.LEFT, padx=(10, 5))
        self.search_entry.bind('<KeyRelease>', self.filter_items)
        
        # Items listbox with scrollbar
        listbox_frame = tk.Frame(item_section, bg=self.bg_color)
        listbox_frame.pack(fill=tk.BOTH, expand=True)
        
        self.items_listbox = tk.Listbox(listbox_frame, height=20, 
                                       selectmode=tk.SINGLE,
                                       bg="white", font=("Arial", 10))
        self.items_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        listbox_scrollbar = tk.Scrollbar(listbox_frame)
        listbox_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.items_listbox.config(yscrollcommand=listbox_scrollbar.set)
        listbox_scrollbar.config(command=self.items_listbox.yview)
        
        # Bind listbox selection
        self.items_listbox.bind('<<ListboxSelect>>', self.on_item_select)
        
        # Item info display
        info_frame = tk.Frame(item_section, bg=self.bg_color)
        info_frame.pack(fill=tk.X, pady=(10, 0))
        
        tk.Label(info_frame, text="Selected Item:", bg=self.bg_color).pack(side=tk.LEFT)
        self.selected_item_label = tk.Label(info_frame, text="None", 
                                           fg="blue", bg=self.bg_color,
                                           font=("Arial", 10, "bold"))
        self.selected_item_label.pack(side=tk.LEFT, padx=(10, 0))
        
        # ========== MIDDLE PANEL ==========
        
        # Sales entry section
        sales_section = tk.LabelFrame(middle_panel, text="Sales Entry", 
                                     font=("Arial", 10, "bold"),
                                     bg=self.bg_color, padx=10, pady=10)
        sales_section.pack(fill=tk.BOTH, expand=True)
        
        # Date entry
        date_frame = tk.Frame(sales_section, bg=self.bg_color)
        date_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(date_frame, text="ቀን:", bg=self.bg_color, width=15, anchor="w").pack(side=tk.LEFT)
        current_date = datetime.now().strftime("%d/%m/%Y")
        self.date_entry = tk.Entry(date_frame, width=25, bg=self.entry_bg)
        self.date_entry.insert(0, current_date)
        self.date_entry.pack(side=tk.LEFT)
        
        tk.Button(date_frame, text="ዛሬ", 
                 command=lambda: self.date_entry.delete(0, tk.END) or 
                                self.date_entry.insert(0, datetime.now().strftime("%d/%m/%Y")),
                 width=8).pack(side=tk.LEFT, padx=(10, 0))
        
        # Item info display (read-only)
        item_info_frame = tk.Frame(sales_section, bg=self.bg_color)
        item_info_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(item_info_frame, text="የዕቃው አይነት:", bg=self.bg_color, 
                width=15, anchor="w").pack(side=tk.LEFT)
        self.item_name_display = tk.Entry(item_info_frame, width=25, 
                                         bg="#e6e6e6", state='readonly')
        self.item_name_display.pack(side=tk.LEFT)
        
        # Quantity entry
        qty_frame = tk.Frame(sales_section, bg=self.bg_color)
        qty_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(qty_frame, text="ብዛት:", bg=self.bg_color, 
                width=15, anchor="w").pack(side=tk.LEFT)
        self.quantity_var = tk.StringVar()
        self.quantity_entry = tk.Entry(qty_frame, textvariable=self.quantity_var,
                                      width=25, bg=self.entry_bg)
        self.quantity_entry.pack(side=tk.LEFT)
        
        # Measurement unit
        unit_frame = tk.Frame(sales_section, bg=self.bg_color)
        unit_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(unit_frame, text="መለኪያ:", bg=self.bg_color, 
                width=15, anchor="w").pack(side=tk.LEFT)
        self.unit_var = tk.StringVar(value="በቁጥር")  # Default unit
        unit_combo = ttk.Combobox(unit_frame, textvariable=self.unit_var,
                                 values=["በቁጥር", "በኪሎ", "በሊትር"], 
                                 state="readonly", width=22)
        unit_combo.pack(side=tk.LEFT)
        
        # Price entry
        price_frame = tk.Frame(sales_section, bg=self.bg_color)
        price_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(price_frame, text="የመሸጫ ዋጋ:", bg=self.bg_color, 
                width=15, anchor="w").pack(side=tk.LEFT)
        self.price_var = tk.StringVar()
        self.price_entry = tk.Entry(price_frame, textvariable=self.price_var,
                                   width=25, bg=self.entry_bg)
        self.price_entry.pack(side=tk.LEFT)
        
        # Total price (auto-calculated)
        total_frame = tk.Frame(sales_section, bg=self.bg_color)
        total_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(total_frame, text="ጠቅላላ ዋጋ:", bg=self.bg_color, 
                width=15, anchor="w").pack(side=tk.LEFT)
        self.total_var = tk.StringVar()
        self.total_label = tk.Label(total_frame, text="0.00", 
                                   bg="white", width=23, anchor="w",
                                   relief="sunken", padx=5)
        self.total_label.pack(side=tk.LEFT)
        
        # Bind quantity and price changes
        self.quantity_var.trace('w', self.calculate_total)
        self.price_var.trace('w', self.calculate_total)
        
        # Add sale button
        add_button_frame = tk.Frame(sales_section, bg=self.bg_color)
        add_button_frame.pack(fill=tk.X, pady=(20, 10))
        
        self.add_sale_btn = tk.Button(add_button_frame, text="ሽያጭ አስገባ", 
                                     command=self.add_sale_record,
                                     bg="#FF5722", fg="white",
                                     font=("Arial", 11, "bold"),
                                     height=2, width=20,
                                     state=tk.DISABLED)
        self.add_sale_btn.pack()
        
        # Recent sales section
        recent_section = tk.LabelFrame(middle_panel, text="የዛሬ ሽያጭ", 
                                      font=("Arial", 10, "bold"),
                                      bg=self.bg_color, padx=10, pady=10)
        recent_section.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        # Today's sales text area
        self.today_sales_text = scrolledtext.ScrolledText(recent_section, 
                                                         height=10,
                                                         bg="white",
                                                         font=("Arial", 9))
        self.today_sales_text.pack(fill=tk.BOTH, expand=True)
        self.today_sales_text.config(state=tk.DISABLED)
        
        # ========== RIGHT PANEL ==========
        
        # PDF Export section
        pdf_section = tk.LabelFrame(right_panel, text="PDF ምድጃ", 
                                   font=("Arial", 10, "bold"),
                                   bg=self.bg_color, padx=10, pady=10)
        pdf_section.pack(fill=tk.BOTH, expand=True)
        
        # PDF Options
        options_frame = tk.Frame(pdf_section, bg=self.bg_color)
        options_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(options_frame, text="ምድጃ አማራጮች:", bg=self.bg_color, 
                font=("Arial", 9, "bold")).pack(anchor="w", pady=(0, 5))
        
        # Date range for export
        date_range_frame = tk.Frame(pdf_section, bg=self.bg_color)
        date_range_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(date_range_frame, text="ከ:", bg=self.bg_color, width=8).pack(side=tk.LEFT)
        self.pdf_from_date = tk.Entry(date_range_frame, width=12, bg=self.entry_bg)
        self.pdf_from_date.insert(0, datetime.now().strftime("%d/%m/%Y"))
        self.pdf_from_date.pack(side=tk.LEFT, padx=(5, 10))
        
        tk.Label(date_range_frame, text="ድረስ:", bg=self.bg_color, width=8).pack(side=tk.LEFT)
        self.pdf_to_date = tk.Entry(date_range_frame, width=12, bg=self.entry_bg)
        self.pdf_to_date.insert(0, datetime.now().strftime("%d/%m/%Y"))
        self.pdf_to_date.pack(side=tk.LEFT, padx=(5, 0))
        
        # Export buttons
        export_btn_frame = tk.Frame(pdf_section, bg=self.bg_color)
        export_btn_frame.pack(fill=tk.X, pady=10)
        
        tk.Button(export_btn_frame, text="የዛሬ ሽያጭ PDF አውጣ", 
                 command=self.export_today_sales_pdf,
                 bg="#9C27B0", fg="white",
                 height=2, width=25).pack(pady=5)
        
        tk.Button(export_btn_frame, text="ሁሉም ሽያጭ PDF አውጣ", 
                 command=self.export_all_sales_pdf,
                 bg="#673AB7", fg="white",
                 height=2, width=25).pack(pady=5)
        
        tk.Button(export_btn_frame, text="በቀን ክልል PDF አውጣ", 
                 command=self.export_date_range_pdf,
                 bg="#3F51B5", fg="white",
                 height=2, width=25).pack(pady=5)
        
        # Summary section
        summary_section = tk.LabelFrame(right_panel, text="የቀኑ ማጠቃለያ", 
                                       font=("Arial", 10, "bold"),
                                       bg=self.bg_color, padx=10, pady=10)
        summary_section.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        # Summary labels
        summary_frame = tk.Frame(summary_section, bg=self.bg_color)
        summary_frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(summary_frame, text="የዛሬ ማጠቃለያ:", bg=self.bg_color,
                font=("Arial", 9, "bold")).pack(anchor="w", pady=(0, 10))
        
        # Summary data
        summary_data_frame = tk.Frame(summary_frame, bg=self.bg_color)
        summary_data_frame.pack(fill=tk.X)
        
        tk.Label(summary_data_frame, text="ጠቅላላ የተሸጡ ዕቃዎች:", bg=self.bg_color,
                width=18, anchor="w").pack(side=tk.LEFT, pady=2)
        self.total_items_label = tk.Label(summary_data_frame, text="0", 
                                         bg="white", width=10, relief="sunken")
        self.total_items_label.pack(side=tk.LEFT, padx=(5, 0), pady=2)
        
        tk.Label(summary_data_frame, text="ጠቅላላ ገቢ:", bg=self.bg_color,
                width=15, anchor="w").pack(side=tk.LEFT, pady=2, padx=(10, 0))
        self.total_revenue_label = tk.Label(summary_data_frame, text="0.00", 
                                           bg="white", width=10, relief="sunken")
        self.total_revenue_label.pack(side=tk.LEFT, padx=(5, 0), pady=2)
        
        # Quick actions
        quick_action_frame = tk.Frame(summary_section, bg=self.bg_color)
        quick_action_frame.pack(fill=tk.X, pady=(20, 0))
        
        tk.Button(quick_action_frame, text="የዛሬ ውሂብ አጽዳ", 
                 command=self.clear_today_data,
                 bg="#FF9800", fg="white",
                 width=20).pack(pady=5)
        
        tk.Button(quick_action_frame, text="ሽያጭ ሪፖርት ተመልከት", 
                 command=self.view_sales_report,
                 bg="#009688", fg="white",
                 width=20).pack(pady=5)
        
        # Status bar
        self.status_bar = tk.Label(self.root, text="ተዘጋጅቷል", 
                                  bd=1, relief=tk.SUNKEN, anchor=tk.W,
                                  bg="#E8E8E8", fg="#333333")
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Initialize daily sales tracking
        self.update_today_sales_display()
    
    def load_excel_file(self):
        """Load existing Excel file"""
        file_path = filedialog.askopenfilename(
            title="Excel ፋይል ይምረጡ",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                self.file_path = file_path
                self.file_label.config(text=os.path.basename(file_path))
                
                # Read the Excel file
                self.df = pd.read_excel(file_path, header=None)
                
                # Extract items from column B (index 1)
                self.extract_items_from_excel()
                
                # Try to get shop owner name from Excel
                self.extract_shop_owner_from_excel()
                
                self.status_bar.config(text=f"ተጭኗል: {os.path.basename(file_path)}")
                messagebox.showinfo("ተሳክቷል", f"Excel ፋይል በተሳካ ሁኔታ ተጭኗል!\n{len(self.items_list)} ዕቃዎች ተገኝተዋል.")
                
            except Exception as e:
                messagebox.showerror("ስህተት", f"Excel ፋይል መጫን አልተሳካም:\n{str(e)}")
    
    def extract_shop_owner_from_excel(self):
        """Extract shop owner name from Excel file"""
        if self.df is None:
            return
        
        try:
            # Look for shop owner in the first few rows
            for i in range(min(10, len(self.df))):
                for j in range(min(15, len(self.df.columns))):
                    cell_value = str(self.df.iloc[i, j])
                    if 'ሰናይት' in cell_value:
                        # Clean up the shop owner name
                        clean_name = cell_value.replace('………………………………………..', '')
                        clean_name = clean_name.strip()
                        if clean_name:
                            self.shop_owner = clean_name
                            self.shop_owner_entry.delete(0, tk.END)
                            self.shop_owner_entry.insert(0, self.shop_owner)
                            return
                    
                    # Also check for shop owner pattern in first row
                    if i == 0 and j == 1:  # Row 1, Column B
                        if cell_value and not any(keyword in cell_value for keyword in ['የ10/18', 'ቀን', 'የዕቃው']):
                            self.shop_owner = cell_value.replace('………………………………………..', '').strip()
                            self.shop_owner_entry.delete(0, tk.END)
                            self.shop_owner_entry.insert(0, self.shop_owner)
                            return
        except:
            pass
    
    def create_new_excel(self):
        """Create a new Excel file with template structure"""
        file_path = filedialog.asksaveasfilename(
            title="አዲስ Excel ፋይል አስቀምጥ",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                # Update shop owner
                self.shop_owner = self.shop_owner_entry.get().strip()
                if not self.shop_owner:
                    self.shop_owner = "ቀጠና 5 ሰናይት"
                
                # Create template structure with shop owner at top
                template_data = {
                    'A': ['', 'ቀን', '', '30/01/2018'],
                    'B': [self.shop_owner + '………………………………………..', 
                         'የዕቃው አይነት', '', 'በቆጠራ የተገኘ'],
                    'C': ['', 'መለኪያ', '', 'በቁጥር'],
                    'D': ['', 'በገቢ/በዝው\nውር/በቆጠራ', '', '29'],
                    'E': ['', 'የተሸጠ\nመጠን', '', ''],
                    'F': ['', 'የግዥ ዋጋ\nያንዱ ዋጋ', '', '43.33'],
                    'G': ['', 'ጠቅላላ ዋጋ', '', '=D4*F4'],
                    'H': ['', 'የመሸጫ ዋጋ\nያንዱ ዋጋ', '', '45'],
                    'I': ['', 'ጠቅላላ ዋጋ', '', '=D4*H4'],
                    'J': ['', 'ቀሪ', '', '=D4-E4'],
                    'K': ['', 'ጉድለት', '', ''],
                    'L': ['', 'ልዩነት\nበገንዘብ', '', ''],
                    'M': ['', 'በትርፍ\nበገንዘብ', '', ''],
                    'N': ['', 'ምርመራ', '', ''],
                    'O': ['', '', '', '1257']  # Last column with total price
                }
                
                df_template = pd.DataFrame(template_data)
                
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df_template.to_excel(writer, index=False, header=False)
                
                self.file_path = file_path
                self.file_label.config(text=os.path.basename(file_path))
                self.df = df_template
                self.items_list = []
                
                self.status_bar.config(text=f"አዲስ Excel ፋይል ተፈጥሯል: {os.path.basename(file_path)}")
                messagebox.showinfo("ተሳክቷል", "አዲስ Excel ፋይል በተሳካ ሁኔታ ተፈጥሯል!\nአሁን ዕቃዎችን እና የሽያጭ መዝገቦችን ማክሰን ይችላሉ.")
                
            except Exception as e:
                messagebox.showerror("ስህተት", f"Excel ፋይል መፍጠር አልተሳካም:\n{str(e)}")
    
    def extract_items_from_excel(self):
        """Extract unique items from the Excel file"""
        if self.df is None:
            return
        
        try:
            # Extract from column B (index 1)
            items_column = self.df.iloc[:, 1]  # Column B
            
            # Get unique values, filter out empty and header rows
            items = []
            for item in items_column.dropna().unique():
                item_str = str(item).strip()
                # Filter out header rows
                if (item_str and 
                    'የ10/18' not in item_str and 
                    'ቀን' not in item_str and
                    'የዕቃው' not in item_str and
                    '………………………………………..' not in item_str):
                    items.append(item_str)
            
            self.items_list = sorted(items)
            self.update_items_listbox()
            
        except Exception as e:
            messagebox.showerror("ስህተት", f"ዕቃዎችን ማውጣት አልተሳካም:\n{str(e)}")
    
    def update_items_listbox(self):
        """Update the items listbox"""
        self.items_listbox.delete(0, tk.END)
        for item in self.items_list:
            self.items_listbox.insert(tk.END, item)
    
    def filter_items(self, event=None):
        """Filter items based on search text"""
        search_text = self.search_var.get().lower()
        
        self.items_listbox.delete(0, tk.END)
        
        if not search_text:
            for item in self.items_list:
                self.items_listbox.insert(tk.END, item)
        else:
            filtered_items = [item for item in self.items_list 
                            if search_text in item.lower()]
            for item in filtered_items:
                self.items_listbox.insert(tk.END, item)
    
    def on_item_select(self, event):
        """Handle item selection from listbox"""
        selection = self.items_listbox.curselection()
        if selection:
            index = selection[0]
            self.current_item = self.items_listbox.get(index)
            self.selected_item_label.config(text=self.current_item)
            self.item_name_display.config(state=tk.NORMAL)
            self.item_name_display.delete(0, tk.END)
            self.item_name_display.insert(0, self.current_item)
            self.item_name_display.config(state='readonly')
            
            # Enable add sale button
            self.add_sale_btn.config(state=tk.NORMAL, bg="#4CAF50")
    
    def calculate_total(self, *args):
        """Calculate total price based on quantity and price"""
        try:
            quantity = float(self.quantity_var.get() or 0)
            price = float(self.price_var.get() or 0)
            total = quantity * price
            self.total_label.config(text=f"{total:.2f}")
        except:
            self.total_label.config(text="0.00")
    
    def add_sale_record(self):
        """Add a new sale record to the Excel file"""
        if not self.current_item or not self.file_path:
            messagebox.showerror("ስህተት", "እባክዎ አንድ ዕቃ ይምረጡ እና Excel ፋይል ይጫኑ!")
            return
        
        # Validate inputs
        try:
            date = self.date_entry.get().strip()
            if not date:
                raise ValueError("ቀን ያስፈልጋል")
            
            quantity = float(self.quantity_var.get().strip())
            if quantity <= 0:
                raise ValueError("ብዛቱ አዎንታዊ መሆን አለበት")
            
            price = float(self.price_var.get().strip())
            if price <= 0:
                raise ValueError("ዋጋው አዎንታዊ መሆን አለበት")
            
            unit = self.unit_var.get()
            
        except ValueError as e:
            messagebox.showerror("ግቤት ስህተት", f"ልክ ያልሆነ ግቤት: {str(e)}")
            return
        
        try:
            # Load the workbook
            wb = load_workbook(self.file_path)
            ws = wb.active
            
            # Update shop owner name in Excel if changed
            self.update_shop_owner_in_excel(wb, ws)
            
            # Find the item row
            item_row = None
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=2).value  # Column B
                if cell_value == self.current_item:
                    item_row = row
                    break
            
            if item_row is None:
                # Item not found, add it
                item_row = self.add_new_item_to_excel(wb, ws)
            
            # Find where to insert new sale (below the item, above next item or empty row)
            insert_row = item_row + 1
            while insert_row <= ws.max_row:
                cell_b = ws.cell(row=insert_row, column=2).value  # Column B
                cell_a = ws.cell(row=insert_row, column=1).value  # Column A
                
                # If we find another item name or the next header section, stop
                if cell_b and (cell_b != self.current_item):
                    break
                
                # If we find an empty row with no data, use it
                if not cell_a and not cell_b:
                    break
                    
                insert_row += 1
            
            # If we reached the end, append new row
            if insert_row > ws.max_row:
                ws.append([None] * 15)  # Add empty row
                insert_row = ws.max_row
            
            # Calculate total price
            total_price = quantity * price
            
            # Insert sale data
            ws.cell(row=insert_row, column=1, value=date)  # Date in column A
            ws.cell(row=insert_row, column=2, value=None)  # Keep column B empty
            ws.cell(row=insert_row, column=3, value=unit)  # Unit in column C
            ws.cell(row=insert_row, column=5, value=quantity)  # Quantity in column E
            ws.cell(row=insert_row, column=8, value=price)  # Price in column H
            ws.cell(row=insert_row, column=9, value=total_price)  # Total in column I (first place)
            ws.cell(row=insert_row, column=15, value=total_price)  # Total in column O (last column - second place)
            
            # Add to today's sales tracking
            self.add_to_today_sales(date, self.current_item, quantity, unit, price, total_price)
            
            # Save the workbook
            wb.save(self.file_path)
            
            # Update the DataFrame
            self.df = pd.read_excel(self.file_path, header=None)
            
            # Clear form
            self.quantity_var.set("")
            self.price_var.set("")
            
            # Update status
            self.status_bar.config(text=f"ሽያጭ ተጨመረ! ቀን: {date}, ዕቃ: {self.current_item}, ጠቅላላ: {total_price:.2f}")
            
            # Show success message
            messagebox.showinfo("ተሳክቷል", 
                              f"የሽያጭ መዝገብ በተሳካ ሁኔታ ተጨመረ!\n\n"
                              f"የንግድ ቦታ: {self.shop_owner}\n"
                              f"ዕቃ: {self.current_item}\n"
                              f"ቀን: {date}\n"
                              f"ብዛት: {quantity} {unit}\n"
                              f"ዋጋ: {price}\n"
                              f"ጠቅላላ: {total_price:.2f}")
            
            # Update today's sales display
            self.update_today_sales_display()
            
        except Exception as e:
            messagebox.showerror("ስህተት", f"የሽያጭ መዝገብ መጨመር አልተሳካም:\n{str(e)}")
            import traceback
            traceback.print_exc()
    
    def add_new_item_to_excel(self, wb, ws):
        """Add a new item to the Excel file"""
        # Find the last row with data
        last_row = ws.max_row
        
        # Add blank line if needed
        if ws.cell(row=last_row, column=1).value or ws.cell(row=last_row, column=2).value:
            ws.append([None] * 15)
            last_row += 1
        
        # Add the item
        item_row = last_row + 1
        ws.cell(row=item_row, column=2, value=self.current_item)  # Item name in column B
        
        return item_row
    
    def update_shop_owner_in_excel(self, wb, ws):
        """Update shop owner name in Excel file"""
        self.shop_owner = self.shop_owner_entry.get().strip()
        
        # Look for shop owner name in first few rows, column B
        for row in range(1, 5):
            cell_value = ws.cell(row=row, column=2).value
            if cell_value and ('………………………………………..' in str(cell_value) or row == 1):
                ws.cell(row=row, column=2, value=self.shop_owner + '………………………………………..')
                return
    
    def add_to_today_sales(self, date, item, quantity, unit, price, total):
        """Add sale to today's tracking"""
        today_date = datetime.now().strftime("%d/%m/%Y")
        if date == today_date:
            self.daily_sales.append({
                'date': date,
                'item': item,
                'quantity': quantity,
                'unit': unit,
                'price': price,
                'total': total
            })
    
    def update_today_sales_display(self):
        """Update today's sales display"""
        self.today_sales_text.config(state=tk.NORMAL)
        self.today_sales_text.delete(1.0, tk.END)
        
        today_date = datetime.now().strftime("%d/%m/%Y")
        today_sales = [s for s in self.daily_sales if s['date'] == today_date]
        
        if today_sales:
            self.today_sales_text.insert(tk.END, f"ሽያጭ ለ {today_date}:\n")
            self.today_sales_text.insert(tk.END, "=" * 60 + "\n")
            
            total_items = 0
            total_revenue = 0
            
            for sale in today_sales:
                self.today_sales_text.insert(tk.END, 
                    f"{sale['item'][:30]:30} | {sale['quantity']:6} {sale['unit'][:8]:8} | {sale['price']:8.2f} | {sale['total']:10.2f}\n")
                total_items += sale['quantity']
                total_revenue += sale['total']
            
            self.today_sales_text.insert(tk.END, "=" * 60 + "\n")
            self.today_sales_text.insert(tk.END, 
                f"ጠቅላላ ዕቃዎች: {total_items:6} | ጠቅላላ ገቢ: {total_revenue:10.2f}\n")
            
            # Update summary labels
            self.total_items_label.config(text=str(total_items))
            self.total_revenue_label.config(text=f"{total_revenue:.2f}")
        else:
            self.today_sales_text.insert(tk.END, f"ለ {today_date} ምንም ሽያጭ አልተመዘገበም")
            self.total_items_label.config(text="0")
            self.total_revenue_label.config(text="0.00")
        
        self.today_sales_text.config(state=tk.DISABLED)
    
    def export_today_sales_pdf(self):
        """Export today's sales to PDF"""
        today_date = datetime.now().strftime("%d/%m/%Y")
        self.export_sales_to_pdf(today_date, today_date, "የዛሬ_ሽያጭ.pdf")
    
    def export_all_sales_pdf(self):
        """Export all sales to PDF"""
        if self.df is None:
            messagebox.showerror("ስህተት", "እባክዎ መጀመሪያ Excel ፋይል ይጫኑ!")
            return
        
        # Get min and max dates from Excel
        dates = []
        for date_val in self.df.iloc[:, 0].dropna():
            try:
                dates.append(str(date_val))
            except:
                pass
        
        if dates:
            from_date = min(dates)
            to_date = max(dates)
            self.pdf_from_date.delete(0, tk.END)
            self.pdf_from_date.insert(0, from_date)
            self.pdf_to_date.delete(0, tk.END)
            self.pdf_to_date.insert(0, to_date)
        
        self.export_sales_to_pdf(self.pdf_from_date.get(), self.pdf_to_date.get(), "ሁሉም_ሽያጭ.pdf")
    
    def export_date_range_pdf(self):
        """Export sales for date range to PDF"""
        from_date = self.pdf_from_date.get().strip()
        to_date = self.pdf_to_date.get().strip()
        
        if not from_date or not to_date:
            messagebox.showerror("ስህተት", "እባክዎ ሁለቱንም የመጀመሪያ እና የመጨረሻ ቀኖች ያስገቡ!")
            return
        
        filename = f"ሽያጭ_{from_date.replace('/', '-')}_ለ_{to_date.replace('/', '-')}.pdf"
        self.export_sales_to_pdf(from_date, to_date, filename)
    
    def export_sales_to_pdf(self, from_date, to_date, filename):
        """Export sales data to PDF"""
        if self.df is None:
            messagebox.showerror("ስህተት", "እባክዎ መጀመሪያ Excel ፋይል ይጫኑ!")
            return
        
        try:
            # Ask for save location
            save_path = filedialog.asksaveasfilename(
                title="PDF ፋይል አስቀምጥ",
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                initialfile=filename
            )
            
            if not save_path:
                return
            
            # Update shop owner
            self.shop_owner = self.shop_owner_entry.get().strip()
            if not self.shop_owner:
                self.shop_owner = "ቀጠና 5 ሰናይት"
            
            # Extract sales data from Excel
            sales_data = self.extract_sales_data(from_date, to_date)
            
            # Create PDF with Amharic font
            doc = SimpleDocTemplate(save_path, pagesize=landscape(A4))
            elements = []
            
            # Register Amharic font
            font_registered = self.register_amharic_font()
            
            # Styles - Use Amharic font if available, otherwise fall back
            styles = getSampleStyleSheet()
            
            if font_registered:
                # Create styles with Amharic font
                title_style = ParagraphStyle(
                    'AmharicTitle',
                    parent=styles['Normal'],
                    fontName='Amharic',
                    fontSize=18,
                    textColor=colors.darkblue,
                    alignment=1,  # Center
                    spaceAfter=12
                )
                
                subtitle_style = ParagraphStyle(
                    'AmharicSubtitle',
                    parent=styles['Normal'],
                    fontName='Amharic',
                    fontSize=14,
                    textColor=colors.darkgreen,
                    alignment=1,
                    spaceAfter=20
                )
                
                normal_style = ParagraphStyle(
                    'AmharicNormal',
                    parent=styles['Normal'],
                    fontName='Amharic',
                    fontSize=11,
                    leading=14
                )
                
                header_style = ParagraphStyle(
                    'AmharicHeader',
                    parent=styles['Normal'],
                    fontName='Amharic',
                    fontSize=11,
                    textColor=colors.white,
                    alignment=1
                )
                
                footer_style = ParagraphStyle(
                    'AmharicFooter',
                    parent=styles['Normal'],
                    fontName='Amharic',
                    fontSize=9,
                    textColor=colors.gray,
                    alignment=1
                )
            else:
                # Fallback styles
                title_style = ParagraphStyle(
                    'Title',
                    parent=styles['Heading1'],
                    fontSize=16,
                    textColor=colors.darkblue,
                    alignment=1,
                    spaceAfter=12
                )
                
                subtitle_style = ParagraphStyle(
                    'Subtitle',
                    parent=styles['Heading2'],
                    fontSize=12,
                    textColor=colors.darkgreen,
                    alignment=1,
                    spaceAfter=20
                )
                
                normal_style = styles['Normal']
                header_style = ParagraphStyle(
                    'Header',
                    parent=styles['Normal'],
                    fontSize=10,
                    textColor=colors.white,
                    alignment=1
                )
                footer_style = styles['Normal']
            
            # Title with shop owner
            elements.append(Paragraph(self.shop_owner, title_style))
            elements.append(Spacer(1, 12))
            
            # Subtitle
            subtitle_text = "የሽያጭ ሪፖርት"
            elements.append(Paragraph(subtitle_text, subtitle_style))
            
            # Date range
            date_range_text = f"ከ: {from_date} - እስከ: {to_date}"
            elements.append(Paragraph(date_range_text, normal_style))
            elements.append(Spacer(1, 15))
            
            # Summary section
            summary_data = self.calculate_summary(sales_data)
            summary_text = f"""
            <b>ማጠቃለያ:</b><br/>
            ጠቅላላ የተሸጡ ዕቃዎች: {summary_data['total_items']}<br/>
            ጠቅላላ ገቢ: {summary_data['total_revenue']:.2f} ብር<br/>
            የሽያጭ ብዛት: {summary_data['num_transactions']}<br/>
            አማካይ በአንድ ሽያጭ: {summary_data['total_revenue']/summary_data['num_transactions'] if summary_data['num_transactions'] > 0 else 0:.2f} ብር
            """
            
            summary_para = Paragraph(summary_text, normal_style)
            elements.append(summary_para)
            elements.append(Spacer(1, 20))
            
            # Create table data with Amharic headers
            if font_registered:
                table_headers = ['ቀን', 'የዕቃው አይነት', 'መለኪያ', 'ብዛት', 'ዋጋ', 'ጠቅላላ']
            else:
                table_headers = ['Date', 'Item', 'Unit', 'Qty', 'Price', 'Total']
            
            table_data = [table_headers]
            
            for sale in sales_data:
                table_data.append([
                    sale.get('date', ''),
                    sale.get('item', '')[:25],  # Limit item name length
                    sale.get('unit', ''),
                    f"{sale.get('quantity', 0):.1f}",
                    f"{sale.get('price', 0):.2f}",
                    f"{sale.get('total', 0):.2f}"
                ])
            
            # Create table
            col_widths = [1.2*inch, 2.8*inch, 0.8*inch, 0.8*inch, 1*inch, 1.2*inch]
            table = Table(table_data, colWidths=col_widths)
            
            # Table style
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E8B57')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Amharic' if font_registered else 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (3, 1), (5, -1), 'RIGHT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Amharic' if font_registered else 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ])
            
            # Add zebra stripes
            for i in range(1, len(table_data)):
                if i % 2 == 0:
                    table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8F8F8'))
            
            table.setStyle(table_style)
            elements.append(table)
            elements.append(Spacer(1, 30))
            
            # Footer with generation date
            gen_date = datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")
            footer_text = f"የተፈጠረው: {gen_date}"
            elements.append(Paragraph(footer_text, footer_style))
            
            # Build PDF
            doc.build(elements)
            
            self.status_bar.config(text=f"PDF ተፈጥሯል: {os.path.basename(save_path)}")
            
            # Show success message in Amharic
            success_msg = f"PDF በተሳካ ሁኔታ ተፈጥሯል!\n\nተፈጥሮአል በ:\n{save_path}"
            messagebox.showinfo("ተሳክቷል", success_msg)
            
            # Option to open the PDF
            if messagebox.askyesno("PDF ክፈት", "PDF ፋይሉን መክፈት ይፈልጋሉ?"):
                try:
                    os.startfile(save_path)
                except:
                    import subprocess
                    subprocess.call(['open', save_path])
            
        except Exception as e:
            messagebox.showerror("ስህተት", f"PDF ማውጣት አልተሳካም:\n{str(e)}")
            import traceback
            traceback.print_exc()
    
    def extract_sales_data(self, from_date, to_date):
        """Extract sales data from Excel for date range"""
        sales_data = []
        
        try:
            # Convert date strings if needed
            from_date_str = str(from_date)
            to_date_str = str(to_date)
            
            for idx, row in self.df.iterrows():
                date_val = str(row[0]) if not pd.isna(row[0]) else ""
                item_val = str(row[1]) if not pd.isna(row[1]) else ""
                unit_val = str(row[2]) if not pd.isna(row[2]) else ""
                qty_val = row[4] if not pd.isna(row[4]) else 0
                price_val = row[7] if not pd.isna(row[7]) else 0
                total_val = row[14] if not pd.isna(row[14]) else 0  # Column O (last column)
                
                # Check if this is a sales record (has date but no item name in column B)
                if date_val and not item_val and qty_val > 0:
                    # Find the item name from above rows
                    item_name = ""
                    for i in range(idx-1, max(0, idx-10), -1):
                        cell_val = str(self.df.iloc[i, 1]) if not pd.isna(self.df.iloc[i, 1]) else ""
                        if cell_val and cell_val in self.items_list:
                            item_name = cell_val
                            break
                    
                    if item_name:
                        # Check if date is within range
                        try:
                            sales_date = date_val
                            # Simple date comparison (assuming DD/MM/YYYY format)
                            if (sales_date >= from_date_str and sales_date <= to_date_str):
                                sales_data.append({
                                    'date': sales_date,
                                    'item': item_name,
                                    'unit': unit_val,
                                    'quantity': float(qty_val),
                                    'price': float(price_val),
                                    'total': float(total_val)
                                })
                        except:
                            # If date comparison fails, include all
                            sales_data.append({
                                'date': date_val,
                                'item': item_name,
                                'unit': unit_val,
                                'quantity': float(qty_val),
                                'price': float(price_val),
                                'total': float(total_val)
                            })
        
        except Exception as e:
            print(f"ሽያጭ ውሂብ ማውጣት ላይ ስህተት: {e}")
        
        return sales_data
    
    def calculate_summary(self, sales_data):
        """Calculate summary from sales data"""
        total_items = sum(sale.get('quantity', 0) for sale in sales_data)
        total_revenue = sum(sale.get('total', 0) for sale in sales_data)
        num_transactions = len(sales_data)
        
        return {
            'total_items': total_items,
            'total_revenue': total_revenue,
            'num_transactions': num_transactions
        }
    
    def clear_today_data(self):
        """Clear today's sales data from memory"""
        today_date = datetime.now().strftime("%d/%m/%Y")
        self.daily_sales = [s for s in self.daily_sales if s['date'] != today_date]
        self.update_today_sales_display()
        messagebox.showinfo("ተጽድቋል", "የዛሬ የሽያጭ ውሂብ ከማህደረ ትውስታ ተጽድቋል.")
    
    def view_sales_report(self):
        """View sales report in a new window"""
        if self.df is None:
            messagebox.showerror("ስህተት", "እባክዎ መጀመሪያ Excel ፋይል ይጫኑ!")
            return
        
        # Create new window
        report_window = tk.Toplevel(self.root)
        report_window.title("የሽያጭ ሪፖርት")
        report_window.geometry("900x600")
        
        # Create text area with scrollbar
        text_frame = tk.Frame(report_window)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        report_text = scrolledtext.ScrolledText(text_frame, wrap=tk.WORD)
        report_text.pack(fill=tk.BOTH, expand=True)
        
        # Generate report
        report_content = self.generate_sales_report()
        report_text.insert(tk.END, report_content)
        report_text.config(state=tk.DISABLED)
        
        # Export button
        export_btn = tk.Button(report_window, text="PDF አውጣ", 
                              command=lambda: self.export_sales_to_pdf(
                                  self.pdf_from_date.get(), 
                                  self.pdf_to_date.get(),
                                  "የሽያጭ_ሪፖርት.pdf"))
        export_btn.pack(pady=10)
    
    def generate_sales_report(self):
        """Generate text sales report"""
        from_date = self.pdf_from_date.get()
        to_date = self.pdf_to_date.get()
        
        sales_data = self.extract_sales_data(from_date, to_date)
        summary = self.calculate_summary(sales_data)
        
        report = f"""
        {'='*70}
        የሽያጭ ሪፖርት
        {'='*70}
        የንግድ ቦታ: {self.shop_owner}
        ጊዜ: {from_date} እስከ {to_date}
        የተፈጠረው: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
        {'='*70}
        
        ማጠቃለያ:
        ጠቅላላ የተሸጡ ዕቃዎች: {summary['total_items']}
        ጠቅላላ ገቢ: {summary['total_revenue']:.2f} ብር
        የሽያጭ ብዛት: {summary['num_transactions']}
        አማካይ በአንድ ሽያጭ: {summary['total_revenue']/summary['num_transactions'] if summary['num_transactions'] > 0 else 0:.2f} ብር
        
        {'='*70}
        ዝርዝር ሽያጭ:
        {'='*70}
        ቀን        | ዕቃ                              | መለኪያ   | ብዛት   | ዋጋ      | ጠቅላላ
        {'-'*85}
        """
        
        for sale in sales_data:
            report += f"{sale['date']:10} | {sale['item']:30} | {sale['unit']:8} | {sale['quantity']:6.1f} | {sale['price']:8.2f} | {sale['total']:10.2f}\n"
        
        report += f"\n{'='*70}\nየሪፖርት መጨረሻ\n{'='*70}"
        
        return report

def main():
    root = tk.Tk()
    app = SalesAutomationSystem(root)
    
    # Center the window
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    root.mainloop()

if __name__ == "__main__":
    # Check and install required packages
    required_packages = ['pandas', 'openpyxl', 'reportlab']
    
    # Check for Amharic font file
    if not os.path.exists("Yebs.ttf"):
        print("""
        ማሳሰቢያ: yebs.ttf ፎንት ፋይል አልተገኘም!
        
        በትክክል አማርኛ ለማሳየት yebs.ttf ፎንት ፋይል ያስፈልጋል።
        
        ለማድረግ:
        1. yebs.ttf ፋይልን ወደ ይህ ማውጫ ይቅዱ
        2. ወይም ፋይሉ በሚገኝበት መንገድ ያረጋግጡ
        
        ያለ አማርኛ ፎንት፣ PDF ሪፖርቶች አማርኛ በትክክል አይታዩም።
        """)
    
    root = tk.Tk()
    app = SalesAutomationSystem(root)
    
    # Center the window
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    root.mainloop()
